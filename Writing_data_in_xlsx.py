# -*- coding: utf-8 -*-
"""
Created on Sun Oct 17 15:33:04 2021

@author: andre
"""


"xlsx file cretaion and naming it"

import xlsxwriter
import pandas as pd
import openpyxl
import time
import requests
import datetime
from datetime import datetime
from datetime import date

"naming the xlsx file by the current date"
today = date.today()
# dd/mm/YY
name = today.strftime("%b-%d-%Y")
wb = xlsxwriter.Workbook('{0}.xlsx'.format(name))
wb.close()

xl = pd.ExcelFile('C:\Trading\Акции доступные в Тинькофф - 2.xlsx')
stocks_in_file = xl.parse('Тикеры')
stocks = []
for i in range(9, len(stocks_in_file), 1):
    company = stocks_in_file.loc[i, 'Ticker']
    stocks.append(company)    

"pandas duplicate table for temporary calculations"
data_table = {stock: pd.DataFrame(columns=('Time', 'Time, s', 'Price', 'Change, %', 'Volume', 'avgVolume3Months', 'avgVolume10Days', 'Bid', 'Ask', 'BidSize', 'AskSize', 'Support price', 'Step price change, %', 'Trend price change, %')) for stock in stocks}

"naming all the sheets of the workbook as tickers"
xl = openpyxl.load_workbook('{0}.xlsx'.format(name))
xl['Sheet1'].title = "{0}".format(stocks[0])

for stock in stocks:    
    ws = xl.create_sheet(stock, 0)
    ws.cell(ws.max_row, ws.max_column).value = 'Time'
    ws.cell(ws.max_row, ws.max_column + 1).value = 'Time, s'
    ws.cell(ws.max_row, ws.max_column + 1).value = 'Price'
    ws.cell(ws.max_row, ws.max_column + 1).value = 'Change, %'
    ws.cell(ws.max_row, ws.max_column + 1).value = 'Volume'
    ws.cell(ws.max_row, ws.max_column + 1).value = 'avgVolume3Months'
    ws.cell(ws.max_row, ws.max_column + 1).value = 'avgVolume10Days'
    ws.cell(ws.max_row, ws.max_column + 1).value = 'Bid'
    ws.cell(ws.max_row, ws.max_column + 1).value = 'Ask'
    ws.cell(ws.max_row, ws.max_column + 1).value = 'BidSize'
    ws.cell(ws.max_row, ws.max_column + 1).value = 'AskSize'
    ws.cell(ws.max_row, ws.max_column + 1).value = 'Support price'
    ws.cell(ws.max_row, ws.max_column + 1).value = 'Step price change, %'
    ws.cell(ws.max_row, ws.max_column + 1).value = 'Trend price change, %'
    ws.cell(ws.max_row, ws.max_column + 1).value = 'Support price time'
    ws.cell(ws.max_row, ws.max_column + 1).value = 'Growth time, h'
    

"writing data in sheets in xlsx"

marketDataUrl = "https://query1.finance.yahoo.com/v7/finance/quote?symbols="
headers = {'User-Agent': 'X'}

class MarketData:
    def __init__(self, symbol, bid, ask):
        self.symbol, self.bid, self.ask = symbol, bid, ask

    def __repr__(self):
        return "[symbol=" + self.symbol + ", bid=" + str(self.bid) + ", ask=" + str(self.ask) + "]" 

def get_market_data(stocks):
    stocks_delimited_by_comma = ",".join(stocks)
    final_url = marketDataUrl + stocks_delimited_by_comma
    response = requests.get(final_url, headers=headers)
    result_list = response.json()["quoteResponse"]["result"]    
    for element in result_list:                
        
        ticker = element["symbol"]
        stock_frame = data_table[ticker] 
        next_row = len(stock_frame) + 1   
        # pandas
        now = datetime.now()                  
        price = element["regularMarketPrice"]
        
        stock_frame.loc[next_row, 'Time'] = now.strftime("%d/%m/%Y %H:%M:%S") 
        current_time = time.time()
        stock_frame.loc[next_row, 'Time, s'] = current_time
        stock_frame.loc[next_row, 'Price'] = price
        stock_frame.loc[next_row, 'Change, %'] = element["regularMarketChangePercent"]   
        stock_frame.loc[next_row, 'Volume'] = element["regularMarketVolume"]
        stock_frame.loc[next_row, 'Bid'] = element["bid"]  
        stock_frame.loc[next_row, 'Ask'] = element["ask"]         
        
        
        if next_row > 1:
            previous_price = stock_frame.loc[next_row - 1, 'Price']                
            previous_price_time = stock_frame.loc[next_row - 1, 'Time, s']    
            
            # find support price
            if stock_frame.loc[next_row - 1, 'Support price'] > price:                
                support_price = price
                support_price_time = stock_frame.loc[next_row, 'Time, s']
            else: 
                support_price = stock_frame.loc[next_row - 1, 'Support price']
                support_price_time = stock_frame.loc[next_row - 1, 'Support price time']
                
            # Step price change, %
            step_price_change = (price - previous_price) * 100 / previous_price
            
            # Trend price change, %
            trend_price_change = (price - support_price) * 100 / support_price

            # Growth time
            growth_time = (current_time - support_price_time) / 3600              
            
        else:
            previous_price = price
            previous_price_time = current_time
            support_price = price
            support_price_time = current_time
            step_price_change = 0
            trend_price_change = 0
                
        stock_frame.loc[next_row, 'Support price'] = support_price
        stock_frame.loc[next_row, 'Support price time'] = support_price_time
        stock_frame.loc[next_row, 'Step price change, %'] = step_price_change
        stock_frame.loc[next_row, 'Trend price change, %'] = trend_price_change
        
        # if growth >= 1 and growth time <= 1 hour, then alert
        if trend_price_change >= 1 and growth_time <= 1:
            bot_message = create_telegram_message(ticker, trend_price_change, growth_time)
            telegram_bot_sendtext(bot_message)
        
        # excel
        sheet = xl[element["symbol"]]    
        next_row_xl = sheet.max_row + 1                             
        sheet.cell(next_row_xl, 1).value = stock_frame.loc[next_row, 'Time']
        sheet.cell(next_row_xl, 2).value = stock_frame.loc[next_row, 'Time, s']
        sheet.cell(next_row_xl, 3).value = price
        sheet.cell(next_row_xl, 4).value = stock_frame.loc[next_row, 'Change, %']
        sheet.cell(next_row_xl, 5).value = stock_frame.loc[next_row, 'Volume']
# =============================================================================
#         sheet.cell(next_row_xl, 6).value = element["averageDailyVolume3Month"]         
#         sheet.cell(next_row_xl, 7).value = element["averageDailyVolume10Day"] 
# =============================================================================
        sheet.cell(next_row_xl, 8).value = stock_frame.loc[next_row, 'Bid']
        sheet.cell(next_row_xl, 9).value = stock_frame.loc[next_row, 'Ask']
# =============================================================================
#         sheet.cell(next_row_xl, 10).value = element["bidSize"]
#         sheet.cell(next_row_xl, 11).value = element["askSize"]
# =============================================================================                
        sheet.cell(next_row_xl, 12).value = support_price        
        sheet.cell(next_row_xl, 13).value = stock_frame.loc[next_row, 'Step price change, %']
        sheet.cell(next_row_xl, 14).value = stock_frame.loc[next_row, 'Trend price change, %']
    return

def telegram_bot_sendtext(bot_message):
   bot_token = '2128623436:AAHatdqcCep5mBthSzLQcCF1vJIBJBswfvw'
   bot_chatID = '267851642'
   send_text = 'https://api.telegram.org/bot' + bot_token + '/sendMessage?chat_id=' + bot_chatID + '&parse_mode=Markdown&text=' + bot_message
   response = requests.get(send_text)
   return response.json()

def create_telegram_message(ticker, trend_price_change, growth_time):
    tv_link = f'https://www.tradingview.com/chart/iNtrciPF/?symbol=NASDAQ%3A{ticker}'
    text = f"""{ticker}: {trend_price_change}% 
    In {growth_time} hours
    {tv_link}"""
    
    return text

"life-cycle of the program"
if __name__ == '__main__':
    while 1 == 1:
        get_market_data(stocks)
        xl.save('{0}.xlsx'.format(name))
        time.sleep(240)
        
